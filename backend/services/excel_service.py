"""
Excel处理服务
"""
import os
import uuid
import json
from typing import List, Dict, Any, Optional
from datetime import date, datetime, time
import decimal
import math
import pandas as pd
from openpyxl import load_workbook
import aiosqlite
from config import UPLOAD_DIR
from database import DATABASE_PATH


class ExcelService:
    """Excel文件处理服务"""

    @staticmethod
    def _json_safe(value: Any) -> Any:
        """
        将 openpyxl/pandas 读出的值转换为可 JSON 序列化的基础类型。
        仅用于保存 sheets 元数据（columns/row_count/preview）到数据库。
        """
        if value is None:
            return None

        # 容错：NaN/Inf
        if isinstance(value, float):
            if math.isnan(value) or math.isinf(value):
                return None
            return value

        # 日期时间
        if isinstance(value, datetime):
            # 统一秒级，避免出现过长小数秒
            try:
                return value.isoformat(sep=' ', timespec='seconds')
            except Exception:
                return value.isoformat()
        if isinstance(value, date) and not isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, time):
            try:
                return value.isoformat(timespec='seconds')
            except Exception:
                return value.isoformat()

        # Decimal 等
        if isinstance(value, decimal.Decimal):
            return float(value)

        # numpy 标量等（如果存在 .item()）
        item = getattr(value, 'item', None)
        if callable(item):
            try:
                return item()
            except Exception:
                pass

        # bytes
        if isinstance(value, (bytes, bytearray)):
            try:
                return value.decode('utf-8', errors='replace')
            except Exception:
                return str(value)

        # 递归结构
        if isinstance(value, (list, tuple)):
            return [ExcelService._json_safe(v) for v in value]
        if isinstance(value, dict):
            return {str(k): ExcelService._json_safe(v) for k, v in value.items()}

        return value
    
    @staticmethod
    def parse_excel(file_path: str) -> Dict[str, Any]:
        """
        解析Excel文件，获取所有Sheet的信息
        
        Returns:
            {
                "sheets": [
                    {
                        "name": "Sheet1",
                        "columns": ["A", "B", "C"],
                        "row_count": 100,
                        "preview": [...前5行数据...]
                    }
                ]
            }
        """
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheets_info = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # 获取列名（第一行）
            columns = []
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if first_row:
                columns = [str(cell) if cell is not None else f"列{i+1}" 
                          for i, cell in enumerate(first_row)]
            
            # 获取行数
            row_count = sheet.max_row - 1 if sheet.max_row else 0
            
            # 获取预览数据（前5行）
            preview = []
            for row in sheet.iter_rows(min_row=2, max_row=6, values_only=True):
                preview.append([ExcelService._json_safe(v) for v in row])
            
            sheets_info.append({
                "name": sheet_name,
                "columns": columns,
                "row_count": row_count,
                "preview": preview
            })
        
        workbook.close()
        return {"sheets": sheets_info}
    
    @staticmethod
    def read_sheet_as_dataframe(file_path: str, sheet_name: str) -> pd.DataFrame:
        """读取指定Sheet为DataFrame"""
        return pd.read_excel(file_path, sheet_name=sheet_name)
    
    @staticmethod
    def read_column(file_path: str, sheet_name: str, column_name: str) -> pd.Series:
        """读取指定列"""
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        if column_name in df.columns:
            return df[column_name]
        raise ValueError(f"列 '{column_name}' 不存在于Sheet '{sheet_name}'")
    
    @staticmethod
    async def save_file_record(file_id: str, filename: str, original_name: str,
                               file_path: str, sheets: List[Dict], owner_id: str) -> None:
        """保存文件记录到数据库"""
        async with aiosqlite.connect(DATABASE_PATH) as db:
            await db.execute(
                """INSERT INTO uploaded_files (id, owner_id, filename, original_name, file_path, sheets)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (file_id, owner_id, filename, original_name, file_path, json.dumps(sheets, ensure_ascii=False))
            )
            await db.commit()
    
    @staticmethod
    async def get_file_record(file_id: str, owner_id: str) -> Optional[Dict]:
        """获取文件记录"""
        async with aiosqlite.connect(DATABASE_PATH) as db:
            db.row_factory = aiosqlite.Row
            cursor = await db.execute(
                "SELECT * FROM uploaded_files WHERE id = ? AND owner_id = ?",
                (file_id, owner_id)
            )
            row = await cursor.fetchone()
            if row:
                return dict(row)
            return None
    
    @staticmethod
    async def delete_file_record(file_id: str, owner_id: str) -> None:
        """从数据库删除文件记录"""
        async with aiosqlite.connect(DATABASE_PATH) as db:
            await db.execute(
                "DELETE FROM uploaded_files WHERE id = ? AND owner_id = ?",
                (file_id, owner_id)
            )
            await db.commit()
    
    @staticmethod
    async def get_all_files(owner_id: str) -> List[Dict]:
        """获取所有上传的文件"""
        async with aiosqlite.connect(DATABASE_PATH) as db:
            db.row_factory = aiosqlite.Row
            cursor = await db.execute(
                "SELECT * FROM uploaded_files WHERE owner_id = ? ORDER BY created_at DESC",
                (owner_id,)
            )
            rows = await cursor.fetchall()
            return [dict(row) for row in rows]

    @staticmethod
    async def get_file_paths(file_ids: List[str], owner_id: str) -> Dict[str, str]:
        """批量获取文件路径映射"""
        if not file_ids:
            return {}
        unique_ids = list(dict.fromkeys([str(fid) for fid in file_ids]))
        placeholders = ", ".join(["?"] * len(unique_ids))
        query = f"SELECT id, file_path FROM uploaded_files WHERE owner_id = ? AND id IN ({placeholders})"
        async with aiosqlite.connect(DATABASE_PATH) as db:
            db.row_factory = aiosqlite.Row
            cursor = await db.execute(query, [owner_id, *unique_ids])
            rows = await cursor.fetchall()
            return {row["id"]: row["file_path"] for row in rows}
    
    @staticmethod
    def export_dataframe(df: pd.DataFrame, output_path: str) -> str:
        """导出DataFrame为Excel文件"""
        df.to_excel(output_path, index=False)
        return output_path
